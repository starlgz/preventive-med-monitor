import openpyxl
import xlrd
import re
from typing import List, Dict, Any, Optional, Tuple
from app.core.logger import logger
from app.parsers.column_mapper import ColumnMapper

class ExcelJobParser:
    """Excel 岗位表形态学解构解析器 2.0
    支持 .xlsx 与 .xls
    特性：
    1. 前 10 行智能表头指纹探测与多级表头自底向上扁平化合成
    2. 跨行跨列合并单元格正规化展开（Span Unfolding）
    3. 脏数据熔断拦截（纯数字、纯序号、合计行自动剔除）
    4. 岗位名称与单位名称智能容错补全
    """

    INVALID_JOB_NAMES = {
        "序号", "编号", "代码", "岗位代码", "职位代码", "合计", "总计", "小计", "说明", "备注",
        "招聘计划", "招聘总表", "岗位表", "一览表", "附表", "附件", "1", "2", "3", "4", "5", "6", "7", "8", "9", "10",
        "1.0", "2.0", "3.0", "4.0", "5.0", "6.0", "7.0", "8.0", "9.0", "10.0"
    }

    @classmethod
    def clean_cell(cls, val: Any) -> str:
        if val is None:
            return ""
        s = str(val).strip()
        # 去除末尾的 .0 （浮点数序号转为整数字符串）
        if re.match(r"^\d+\.0$", s):
            s = s[:-2]
        return s

    @classmethod
    def parse_file(cls, file_path: str, default_unit_name: Optional[str] = None) -> List[Dict[str, Any]]:
        if file_path.endswith(".xlsx"):
            return cls._parse_xlsx(file_path, default_unit_name)
        elif file_path.endswith(".xls"):
            return cls._parse_xls(file_path, default_unit_name)
        else:
            logger.warning(f"Unsupported excel format: {file_path}")
            return []

    @classmethod
    def _detect_header(cls, rows_data: List[List[str]]) -> Tuple[int, int, Dict[str, int]]:
        """智能探测表头起始行、结束行及映射字典
        返回：(header_start_row, header_end_row, col_map)
        """
        max_scan = min(12, len(rows_data))
        best_score = 0
        best_single_idx = -1

        # 先找单行得分最高者
        for r_idx in range(max_scan):
            score = ColumnMapper.evaluate_header_score(rows_data[r_idx])
            if score > best_score:
                best_score = score
                best_single_idx = r_idx

        if best_single_idx == -1:
            return -1, -1, {}

        # 尝试合成多级表头 (如果上一行或下一行也有表头特征)
        header_rows = [rows_data[best_single_idx]]
        start_row = best_single_idx
        end_row = best_single_idx

        # 检查上一行是否为复合大类标题（如包含“资格条件”、“学历学位”）
        if best_single_idx > 0:
            prev_row = rows_data[best_single_idx - 1]
            if any(k in "".join(prev_row) for k in ["资格条件", "招聘条件", "基本条件", "学历学位", "考试科目"]):
                header_rows.insert(0, prev_row)
                start_row = best_single_idx - 1

        # 检查下一行是否为二级拆分（如上一行是“资格条件”，下一行是“学历”、“专业”）
        if best_single_idx + 1 < len(rows_data):
            next_row = rows_data[best_single_idx + 1]
            if ColumnMapper.evaluate_header_score(next_row) >= 4:
                header_rows.append(next_row)
                end_row = best_single_idx + 1

        synthesized_headers = ColumnMapper.synthesize_multi_headers(header_rows)
        col_map = ColumnMapper.map_columns(synthesized_headers)

        # 如果合成表头未匹配到 job_name，再退回单行匹配
        if "job_name" not in col_map and "major" not in col_map:
            col_map = ColumnMapper.map_columns(rows_data[best_single_idx])
            start_row = best_single_idx
            end_row = best_single_idx

        return start_row, end_row, col_map

    @classmethod
    def _is_invalid_job_name(cls, name: str) -> bool:
        """判定岗位名称是否为脏数据/序号"""
        if not name:
            return True
        name_clean = name.strip()
        if name_clean in cls.INVALID_JOB_NAMES:
            return True
        # 纯数字
        if re.match(r"^\d+$", name_clean):
            return True
        # 纯序号如 01、001、A01
        if re.match(r"^[A-Za-z]?\d{1,4}$", name_clean) and len(name_clean) <= 4:
            return True
        return False

    @classmethod
    def _parse_xlsx(cls, file_path: str, default_unit_name: Optional[str]) -> List[Dict[str, Any]]:
        jobs = []
        try:
            wb = openpyxl.load_workbook(file_path, data_only=True)
            for sheet_name in wb.sheetnames:
                ws = wb[sheet_name]
                sheet_jobs = cls._parse_xlsx_sheet(ws, default_unit_name)
                jobs.extend(sheet_jobs)
            wb.close()
        except Exception as e:
            logger.error(f"Error parsing xlsx file {file_path}: {e}")
        return jobs

    @classmethod
    def _parse_xlsx_sheet(cls, ws, default_unit_name: Optional[str]) -> List[Dict[str, Any]]:
        raw_rows = []
        for row in ws.iter_rows(values_only=False):
            raw_rows.append(row)

        if not raw_rows:
            return []

        # 提取单元格文本矩阵
        rows_text = []
        for row in raw_rows:
            rows_text.append([cls.clean_cell(cell.value) for cell in row])

        # 探测表头
        start_row, end_row, col_map = cls._detect_header(rows_text)
        if "job_name" not in col_map and "major" not in col_map:
            return []

        # 展开合并单元格映射
        merged_ranges = list(ws.merged_cells.ranges)
        jobs = []
        last_unit = default_unit_name or ""

        for r_idx in range(end_row + 1, len(raw_rows)):
            row = raw_rows[r_idx]
            clean_row = []
            for c_idx, cell in enumerate(row):
                actual_val = cell.value
                for m_range in merged_ranges:
                    if (r_idx + 1 >= m_range.min_row and r_idx + 1 <= m_range.max_row and
                        c_idx + 1 >= m_range.min_col and c_idx + 1 <= m_range.max_col):
                        actual_val = ws.cell(row=m_range.min_row, column=m_range.min_col).value
                        break
                clean_row.append(cls.clean_cell(actual_val))

            if not any(clean_row):
                continue

            job = cls._extract_job_dict(clean_row, col_map, last_unit, default_unit_name)
            if job:
                if job["unit_name"] and job["unit_name"] != "未指定招聘单位":
                    last_unit = job["unit_name"]
                jobs.append(job)

        return jobs

    @classmethod
    def _parse_xls(cls, file_path: str, default_unit_name: Optional[str]) -> List[Dict[str, Any]]:
        jobs = []
        try:
            book = xlrd.open_workbook(file_path, formatting_info=True)
            for sheet_idx in range(book.nsheets):
                sheet = book.sheet_by_index(sheet_idx)
                sheet_jobs = cls._parse_xls_sheet(sheet, default_unit_name)
                jobs.extend(sheet_jobs)
        except Exception:
            # 部分 xls 库如果 formatting_info 不支持则降级读取
            try:
                book = xlrd.open_workbook(file_path, formatting_info=False)
                for sheet_idx in range(book.nsheets):
                    sheet = book.sheet_by_index(sheet_idx)
                    sheet_jobs = cls._parse_xls_sheet(sheet, default_unit_name)
                    jobs.extend(sheet_jobs)
            except Exception as e2:
                logger.error(f"Error parsing xls file {file_path}: {e2}")
        return jobs

    @classmethod
    def _parse_xls_sheet(cls, sheet, default_unit_name: Optional[str]) -> List[Dict[str, Any]]:
        if sheet.nrows == 0:
            return []

        # 提取单元格文本矩阵
        rows_text = []
        for r in range(sheet.nrows):
            rows_text.append([cls.clean_cell(sheet.cell_value(r, c)) for c in range(sheet.ncols)])

        start_row, end_row, col_map = cls._detect_header(rows_text)
        if "job_name" not in col_map and "major" not in col_map:
            return []

        # 合并单元格展开
        merged_cells = sheet.merged_cells if hasattr(sheet, "merged_cells") else []
        jobs = []
        last_unit = default_unit_name or ""

        for r_idx in range(end_row + 1, sheet.nrows):
            clean_row = []
            for c_idx in range(sheet.ncols):
                val = sheet.cell_value(r_idx, c_idx)
                for rlow, rhigh, clow, chigh in merged_cells:
                    if rlow <= r_idx < rhigh and clow <= c_idx < chigh:
                        val = sheet.cell_value(rlow, clow)
                        break
                clean_row.append(cls.clean_cell(val))

            if not any(clean_row):
                continue

            job = cls._extract_job_dict(clean_row, col_map, last_unit, default_unit_name)
            if job:
                if job["unit_name"] and job["unit_name"] != "未指定招聘单位":
                    last_unit = job["unit_name"]
                jobs.append(job)

        return jobs

    @classmethod
    def _extract_job_dict(cls, row: List[str], col_map: Dict[str, int], last_unit: str, default_unit_name: Optional[str]) -> Optional[Dict[str, Any]]:
        """从解构后的行数据中抽取并清洗岗位实体"""
        def get_val(key: str) -> str:
            if key in col_map and col_map[key] < len(row):
                return row[col_map[key]].strip()
            return ""

        unit = get_val("unit_name")
        job_name = get_val("job_name")
        major_raw = get_val("major")
        job_code = get_val("job_code")
        headcount_raw = get_val("headcount")
        education = get_val("education")
        degree = get_val("degree")
        age = get_val("age")
        employment_type = get_val("employment_type")
        other_req = get_val("other_requirements")

        # 1. 脏数据拦截与熔断
        if cls._is_invalid_job_name(job_name):
            # 如果 job_name 是纯序号或无效字符，尝试从 major 或 job_code 中寻找线索，否则忽略此行
            if not major_raw and not other_req:
                return None
            job_name = ""

        # 2. 如果缺少 job_name 但有明确的专业要求
        if not job_name:
            if "医师" in major_raw or "公卫" in major_raw or "预防" in major_raw:
                job_name = "专职技术岗"
            elif other_req and ("岗" in other_req or "职位" in other_req):
                job_name = "专业技术岗"
            else:
                job_name = "招聘岗位"

        # 3. 单位名称确定
        if not unit:
            unit = last_unit or default_unit_name or "未指定招聘单位"

        # 4. 人数解析正规化
        headcount = 1
        if headcount_raw:
            digits = re.findall(r"\d+", headcount_raw)
            if digits:
                try:
                    headcount = int(digits[0])
                except:
                    headcount = 1

        # 5. 合并学历与学位
        full_edu = education
        if degree and degree not in full_edu:
            full_edu = f"{full_edu}（{degree}）" if full_edu else degree

        return {
            "unit_name": unit,
            "job_name": job_name,
            "job_code": job_code,
            "headcount": headcount,
            "education": full_edu,
            "major_raw": major_raw,
            "age": age,
            "employment_type": employment_type,
            "other_requirements": other_req
        }
