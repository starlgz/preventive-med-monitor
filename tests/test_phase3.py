import os
import openpyxl
import docx
from app.parsers.column_mapper import ColumnMapper
from app.parsers.excel_parser import ExcelJobParser
from app.parsers.word_parser import WordJobParser
from app.parsers.html_table_parser import HtmlTableJobParser
from app.parsers.dispatcher import ParserDispatcher

def create_mock_excel(path: str):
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "岗位表"
    
    # 写入表头
    ws.append(["用人单位", "岗位代码", "招聘岗位", "招聘人数", "专业要求", "学历要求", "备注"])
    
    # 写入数据行 (包含预防医学与临床医学)
    ws.append(["杭州市疾病预防控制中心", "JK001", "突发事件处置岗", 2, "预防医学、公共卫生", "本科及以上", "在编"])
    ws.append(["杭州市疾病预防控制中心", "JK002", "慢性病防治岗", 1, "流行病与卫生统计学", "硕士研究生", "在编"])
    ws.append(["杭州市第一人民医院", "YY001", "临床医师", 3, "临床医学", "本科及以上", "合同制"])
    
    os.makedirs(os.path.dirname(path), exist_ok=True)
    wb.save(path)

def create_mock_word(path: str):
    doc = docx.Document()
    table = doc.add_table(rows=2, cols=6)
    headers = ["招聘单位", "岗位名称", "招聘人数", "专业要求", "学历要求", "备注"]
    for col_idx, h in enumerate(headers):
        table.cell(0, col_idx).text = h
        
    data = ["宁波市疾控中心", "理化检验岗", "1", "卫生检验与检疫、预防医学", "本科", "全额拨款事业编"]
    for col_idx, d in enumerate(data):
        table.cell(1, col_idx).text = d
        
    os.makedirs(os.path.dirname(path), exist_ok=True)
    doc.save(path)

def run_tests():
    excel_path = "data/attachments/test_mock_job.xlsx"
    word_path = "data/attachments/test_mock_job.docx"
    create_mock_excel(excel_path)
    create_mock_word(word_path)
    
    print("=== 1. 测试 Excel 岗位表解析 ===")
    excel_jobs = ExcelJobParser.parse_file(excel_path)
    print(f"Excel 提取到岗位数量: {len(excel_jobs)}")
    for j in excel_jobs:
        print(f"  - [{j['unit_name']}] 岗位:{j['job_name']}, 专业:{j['major_req']}, 人数:{j['headcount']}, 学历:{j['education']}")
        
    print("\n=== 2. 测试 Word 岗位表解析 ===")
    word_jobs = WordJobParser.parse_file(word_path)
    print(f"Word 提取到岗位数量: {len(word_jobs)}")
    for j in word_jobs:
        print(f"  - [{j['unit_name']}] 岗位:{j['job_name']}, 专业:{j['major_req']}, 人数:{j['headcount']}, 学历:{j['education']}")

    print("\n=== 3. 测试 HTML 内嵌表格解析 ===")
    html_sample = """
    <html>
      <body>
        <h1>2026年温州市疾控中心招聘公告</h1>
        <table border="1">
          <tr><th>单位名称</th><th>招聘岗位</th><th>专业要求</th><th>招聘人数</th></tr>
          <tr><td>温州市疾病预防控制中心</td><td>公共卫生医师</td><td>预防医学、公共卫生</td><td>2</td></tr>
        </table>
      </body>
    </html>
    """
    html_jobs = HtmlTableJobParser.parse_html_tables(html_sample, default_unit="温州市疾控中心")
    print(f"HTML 提取到岗位数量: {len(html_jobs)}")
    for j in html_jobs:
        print(f"  - [{j['unit_name']}] 岗位:{j['job_name']}, 专业:{j['major_req']}, 人数:{j['headcount']}")

if __name__ == "__main__":
    run_tests()
